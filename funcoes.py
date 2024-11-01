import os
import pandas as pd
from openpyxl.styles import PatternFill

def registrar_nota():
    os.system('clear')
    nome = input('Digite o nome do aluno: ')
    notas = []
    for n in range(1, 5):
        nota = float(input('Digite a nota do aluno[-1 --> Sair]: '))
        #Fazer verificação pra nota não ser maior que 20. Caso seja, Acuse e renicie o loop
        notas.append(nota)
    
    media = sum(notas) / len(notas)


    return {'Nome': nome, 'Nota': media}

def media(alunos):
    #Colocar notas em uma lista para colocar na tabela excel
    os.system('clear')
    alunos_aprovados = []
    
    alunos_recuperacao = []
    alunos_reprovados = []
    for aluno in alunos:
        print(aluno)
        if 5 < aluno['Nota'] < 7:
            alunos_recuperacao.append(aluno)
        elif aluno['Nota'] >= 7:
            alunos_aprovados.append(aluno)
        else:
            alunos_reprovados.append(aluno)
                
    return alunos_aprovados, alunos_recuperacao, alunos_reprovados


def atualizar_nota():
    #Desenvolver isso aqui
    pass

def exportar_excel(aprovados, recuperacao, reprovados):
    #colocar as notas inves das situações
    dados_aprovados = {
        'Nome': [aluno['Nome'] for aluno in aprovados],
        'Nota': [aluno['Nota'] for aluno in aprovados],
        'Situação': 'Aprovado'
    }
    dados_recuperacao = {
        'Nome': [aluno['Nome'] for aluno in recuperacao],
        'Nota': [aluno['Nota'] for aluno in recuperacao],
        'Situação': 'Recuperação'
    }
    dados_reprovados = {
        'Nome': [aluno['Nome'] for aluno in reprovados],
        'Nota': [aluno['Nota'] for aluno in reprovados],
        'Situação': 'Reprovado'
    }
    
    dataframeApr = pd.DataFrame(dados_aprovados)
    dataframeRec = pd.DataFrame(dados_recuperacao)
    dataframeRpv = pd.DataFrame(dados_reprovados)

    with pd.ExcelWriter('Alunos.xlsx', engine='openpyxl') as writer:
        dataframeApr.to_excel(writer, sheet_name='Aprovados', index=False)
        dataframeRec.to_excel(writer, sheet_name='Recuperação', index=False)
        dataframeRpv.to_excel(writer, sheet_name='Reprovados', index=False)
        
        workbook = writer.book
        worksheetApr = writer.sheets['Aprovados']
        worksheetRec = writer.sheets['Recuperação']
        worksheetRpv = writer.sheets['Reprovados']

        verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        vermelho = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for row in range(2, len(dataframeApr) + 2):
            worksheetApr.cell(row=row, column=3).fill = verde

        for row in range(2, len(dataframeRec) + 2):
            worksheetRec.cell(row=row, column=3).fill = amarelo    
    
        for row in range(2, len(dataframeRec) + 2):
            worksheetRpv.cell(row=row, column=3).fill = vermelho